#!/usr/bin/env python3
"""
generate_idp.py
为指定人员（或指定盘点的全部关键人才）生成 IDP 培养方案草稿，输出为 Excel。

用法：
  # 为某次盘点所有第一/第二梯队生成IDP
  python3 skills/idp-design/scripts/generate_idp.py \
    --inventory 2025H1 \
    --tiers 1,2 \
    --output artifacts/2025H1-IDP草稿.xlsx

  # 为单人生成IDP（手动触发）
  python3 skills/idp-design/scripts/generate_idp.py \
    --name 张三 \
    --inventory 2025H1 \
    --target_level 二级管理者 \
    --mentor 李四 \
    --output artifacts/张三-IDP.xlsx

输出 Excel 包含：
  Sheet1: IDP汇总（每人一行，含培养目标/短板/导师/任务数/完成率）
  Sheet2: IDP详情（每条任务一行，70-20-10分类，含截止日期/负责人/验收标准）
"""
import argparse, json, sys
from pathlib import Path
from datetime import datetime, timedelta
from copy import deepcopy

BASE = Path(__file__).resolve().parent.parent.parent.parent
PROFILE_FILE = BASE / "data" / "talent_profiles.json"
MENTOR_FILE = Path(__file__).resolve().parent.parent / "data" / "mentor_pool.json"
FIXED_FILE = Path(__file__).resolve().parent.parent / "data" / "fixed_programs.json"

# 唐久胜任力模型——各维度发展建议（70-20-10框架）
# 70=工作历练，20=辅导辅助，10=系统学习
COMPETENCY_DEVELOP_GUIDE = {
    "专业": {
        "70": [
            {"task": "主导一次跨部门质量改善项目，输出改善报告", "months": 3, "kpi": "项目验收通过，改善效果可量化"},
            {"task": "承担本部门年度专项课题研究，形成报告并汇报", "months": 4, "kpi": "汇报获管理层认可"},
        ],
        "20": [
            {"task": "参加行业标杆企业交流学习，输出对标报告", "months": 2, "kpi": "提交≥1份可操作对标报告"},
            {"task": "跟随导师参与战略分析会议，进行复盘总结", "months": 2, "kpi": "完成≥2次复盘记录"},
        ],
        "10": [
            {"task": "完成《行业洞察与战略思维》课程（内部/外部均可）", "months": 1, "kpi": "获得结业证书或完成测评"},
        ],
    },
    "诚信": {
        "70": [
            {"task": "负责一项高透明度的质量管控项目，全程留痕可追溯", "months": 3, "kpi": "无质量数据篡改记录"},
            {"task": "主持本部门廉洁文化宣贯活动", "months": 1, "kpi": "完成活动并有会议记录"},
        ],
        "20": [
            {"task": "与导师进行价值观对齐深度谈话（≥2次）", "months": 2, "kpi": "提交反思日志"},
        ],
        "10": [
            {"task": "完成《职业道德与廉洁从业》线上课程", "months": 1, "kpi": "完成学习并通过测评"},
        ],
    },
    "友好": {
        "70": [
            {"task": "主导一次跨部门协作项目，担任项目协调人", "months": 3, "kpi": "项目参与方满意度≥80%"},
            {"task": "在团队内推行每月1次正式复盘会，持续3个月", "months": 3, "kpi": "完成3次并有会议纪要"},
        ],
        "20": [
            {"task": "导师辅导：学习冲突处理与团队激励技巧（≥3次)", "months": 2, "kpi": "完成辅导并提交学习记录"},
        ],
        "10": [
            {"task": "完成《教练型领导力》或《高效团队协作》课程", "months": 1, "kpi": "结业证书或完成测评"},
        ],
    },
    "利他": {
        "70": [
            {"task": "担任新员工导师，辅导≥1名新员工完成试用期考核", "months": 3, "kpi": "新员工试用期通过率100%"},
            {"task": "参与客户满意度改善项目，负责客户需求收集与反馈", "months": 3, "kpi": "参与客户调研≥5次，形成改善方案"},
        ],
        "20": [
            {"task": "跟随导师参与客户拜访或服务改善复盘", "months": 2, "kpi": "完成≥3次陪访并输出总结"},
        ],
        "10": [
            {"task": "完成《以客户为中心》或《人才发展》专题课程", "months": 1, "kpi": "完成学习并通过测评"},
        ],
    },
    "勤奋": {
        "70": [
            {"task": "承担一个高挑战性、强度大的临时专项任务", "months": 2, "kpi": "按时完成且质量达标"},
            {"task": "持续3个月每周产出工作日志，记录关键问题与行动", "months": 3, "kpi": "日志完整率≥90%"},
        ],
        "20": [
            {"task": "与导师进行目标管理与自律习惯辅导（≥2次）", "months": 2, "kpi": "完成辅导记录"},
        ],
        "10": [
            {"task": "完成《时间管理与高效执行》课程", "months": 1, "kpi": "完成学习"},
        ],
    },
    "责任": {
        "70": [
            {"task": "独立负责一个完整项目（从立项到交付），承担结果", "months": 4, "kpi": "项目按时交付，KPI达标"},
            {"task": "在现岗位上牵头解决一个遗留问题，形成结案报告", "months": 2, "kpi": "问题解决并有验收记录"},
        ],
        "20": [
            {"task": "与导师定期复盘：在哪些事上表现出了担当（≥3次）", "months": 3, "kpi": "提交3份复盘记录"},
        ],
        "10": [
            {"task": "完成《结果导向与责任文化》课程", "months": 1, "kpi": "完成学习"},
        ],
    },
    "坚持": {
        "70": [
            {"task": "设定一个6个月长期目标，每月汇报进展并坚持完成", "months": 6, "kpi": "最终达成目标"},
            {"task": "在遭遇阻力时主导推进一项有难度的改善项目", "months": 3, "kpi": "项目推进未中断，形成结果"},
        ],
        "20": [
            {"task": "与导师分享逆境经历，提炼韧性策略（≥2次深度谈话）", "months": 2, "kpi": "提交个人成长反思"},
        ],
        "10": [
            {"task": "完成《逆商与韧性领导》课程", "months": 1, "kpi": "完成学习"},
        ],
    },
}

# 每个梯队的默认培养目标
TIER_GOAL = {
    1: "12-18个月内具备晋升至上一层级管理岗的能力",
    2: "18-24个月内成为专业骨干/后备管理者",
    3: "24-36个月内提升专业能力，稳定现岗绩效",
}


def load_json(path):
    if path.exists():
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_profiles(data):
    with open(PROFILE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def find_mentor(name, mentor_pool, expertise_dims):
    """从导师库里找最匹配的导师（擅长领域与短板维度重叠最多）"""
    if name:  # 手动指定
        for m in mentor_pool.get("mentors", []):
            if m["name"] == name:
                return m
        return {"name": name, "dept": "", "level": "", "expertise": []}
    # 自动匹配
    best, best_score = None, -1
    for m in mentor_pool.get("mentors", []):
        score = len(set(m.get("expertise", [])) & set(expertise_dims))
        if score > best_score:
            best, best_score = m, score
    return best


def generate_tasks_for_person(person_data, inventory_snapshot, fixed_programs, mentor_name=None, target_level=None):
    """为单人生成任务列表"""
    dim_scores = inventory_snapshot.get("dim_scores", {})
    tier = inventory_snapshot.get("tier", 2)
    if target_level is None:
        target_level = inventory_snapshot.get("target_level", "")

    # 找出最弱的3个维度（有分数的）
    valid_dims = {k: v for k, v in dim_scores.items() if v is not None}
    weak_dims = sorted(valid_dims, key=lambda d: valid_dims[d])[:3] if valid_dims else list(COMPETENCY_DEVELOP_GUIDE.keys())[:2]

    tasks = []
    today = datetime.today()

    # 针对短板维度生成发展任务（每个维度取70和10各1条）
    for dim in weak_dims:
        guide = COMPETENCY_DEVELOP_GUIDE.get(dim, {})
        for category, items in [("70", guide.get("70", [])), ("20", guide.get("20", [])), ("10", guide.get("10", []))]:
            if not items:
                continue
            item = items[0]  # 取第一条最典型任务
            due = today + timedelta(days=30 * item["months"])
            tasks.append({
                "dim": dim,
                "category": category,
                "category_label": {"70": "工作历练(70%)", "20": "辅导辅助(20%)", "10": "系统学习(10%)"}[category],
                "task": item["task"],
                "kpi": item["kpi"],
                "months": item["months"],
                "due_date": due.strftime("%Y-%m-%d"),
                "status": "未开始",
                "progress_note": "",
                "source": "competency"
            })

    # 加入固定培训项目（所有关键人才必须完成）
    for prog in fixed_programs.get("programs", []):
        months = prog.get("months", 1)
        due = today + timedelta(days=30 * months)
        tasks.append({
            "dim": prog.get("competency_dim", "通用"),
            "category": prog.get("category", "10"),
            "category_label": prog.get("category_label", "系统学习(10%)"),
            "task": prog.get("name", ""),
            "kpi": prog.get("kpi", "完成并通过测评"),
            "months": months,
            "due_date": due.strftime("%Y-%m-%d"),
            "status": "未开始",
            "progress_note": "",
            "source": "fixed_program"
        })

    return tasks, weak_dims


def write_excel(all_idps, output_path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(json.dumps({"status": "error", "message": "缺少 openpyxl"}, ensure_ascii=False))
        sys.exit(1)

    wb = openpyxl.Workbook()

    # ── Sheet1: IDP汇总 ──
    ws1 = wb.active
    ws1.title = "IDP汇总"

    header1 = ["员工姓名", "部门", "梯队", "培养目标", "素质短板维度", "导师", "任务总数", "固定项目数", "完成率", "IDP生成时间"]
    COLORS = {"header": "1F497D", "tier1": "DAEEF3", "tier2": "EBF1DE", "tier3": "FFF2CC"}

    def hdr_cell(ws, row, col, val, bg="1F497D"):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return c

    def data_cell(ws, row, col, val, bg=None):
        c = ws.cell(row=row, column=col, value=val)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        return c

    for ci, h in enumerate(header1, 1):
        hdr_cell(ws1, 1, ci, h)

    tier_bg = {1: COLORS["tier1"], 2: COLORS["tier2"], 3: COLORS["tier3"]}

    for ri, idp in enumerate(all_idps, 2):
        bg = tier_bg.get(idp["tier"], "FFFFFF")
        vals = [
            idp["name"], idp["dept"], f"第{idp['tier']}梯队",
            idp["goal"], "、".join(idp["weak_dims"]),
            idp["mentor_name"],
            idp["task_count"], idp["fixed_count"], "0%",
            idp["generated_at"]
        ]
        for ci, v in enumerate(vals, 1):
            data_cell(ws1, ri, ci, v, bg)

    for ci in range(1, len(header1) + 1):
        ws1.column_dimensions[get_column_letter(ci)].width = [12, 10, 8, 30, 20, 10, 8, 8, 8, 18][ci - 1]
    ws1.row_dimensions[1].height = 25

    # ── Sheet2: IDP详情 ──
    ws2 = wb.create_sheet("IDP任务明细")
    header2 = ["员工姓名", "部门", "梯队", "发展维度", "任务分类", "培养任务", "验收标准", "周期(月)", "截止日期", "来源", "状态", "进展备注"]
    for ci, h in enumerate(header2, 1):
        hdr_cell(ws2, 1, ci, h)

    row2 = 2
    for idp in all_idps:
        bg = tier_bg.get(idp["tier"], "FFFFFF")
        for task in idp["tasks"]:
            vals = [
                idp["name"], idp["dept"], f"第{idp['tier']}梯队",
                task["dim"], task["category_label"],
                task["task"], task["kpi"],
                task["months"], task["due_date"],
                "固定项目" if task["source"] == "fixed_program" else "素质发展",
                task["status"], task["progress_note"]
            ]
            for ci, v in enumerate(vals, 1):
                data_cell(ws2, row2, ci, v, bg)
            row2 += 1

    col_widths2 = [12, 10, 8, 8, 14, 40, 30, 8, 12, 10, 8, 20]
    for ci, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[1].height = 25

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--inventory", required=True)
    parser.add_argument("--tiers", default="1,2", help="生成IDP的梯队，逗号分隔，如1,2")
    parser.add_argument("--name", help="单人模式：指定员工姓名")
    parser.add_argument("--target_level", help="单人模式：目标层级")
    parser.add_argument("--mentor", help="单人模式：指定导师姓名")
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    profiles = load_json(PROFILE_FILE)
    mentor_pool = load_json(MENTOR_FILE)
    fixed_programs = load_json(FIXED_FILE)

    if not profiles:
        print(json.dumps({"status": "error", "message": "人才档案为空，请先运行 import_calibrated.py"}, ensure_ascii=False))
        sys.exit(1)

    target_tiers = [int(t.strip()) for t in args.tiers.split(",")]
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    # 筛选目标人员
    # 单人模式（--name 指定）：忽略梯队过滤，任何梯队都可生成IDP
    single_mode = bool(args.name)
    target_people = []
    for name, p in profiles.get("talents", {}).items():
        if args.name and name != args.name:
            continue
        inv_snap = p.get("inventories", {}).get(args.inventory)
        if not inv_snap:
            # 单人模式下即使没有盘点快照也允许生成（用空快照）
            if single_mode:
                inv_snap = {"tier": 2, "dim_scores": {}, "perf_tier": "", "target_level": args.target_level or ""}
            else:
                continue
        tier = inv_snap.get("tier", -1)
        if not single_mode and tier not in target_tiers:
            continue
        target_people.append((name, p, inv_snap))

    if not target_people:
        print(json.dumps({"status": "error", "message": f"未找到符合条件的人员（盘点={args.inventory}，梯队={target_tiers}）"}, ensure_ascii=False))
        sys.exit(1)

    all_idps = []
    for name, person, inv_snap in target_people:
        tier = inv_snap.get("tier", 2)
        tasks, weak_dims = generate_tasks_for_person(
            person, inv_snap, fixed_programs,
            mentor_name=args.mentor if args.name else None,
            target_level=args.target_level
        )
        mentor = find_mentor(
            args.mentor if args.name else None,
            mentor_pool,
            weak_dims
        )
        mentor_name = mentor["name"] if mentor else "待分配"
        goal = TIER_GOAL.get(tier, "提升综合能力")
        if args.target_level:
            goal = f"目标：{args.target_level}，{goal}"

        fixed_count = sum(1 for t in tasks if t["source"] == "fixed_program")

        idp_record = {
            "name": name,
            "dept": person.get("dept", ""),
            "tier": tier,
            "goal": goal,
            "weak_dims": weak_dims,
            "mentor_name": mentor_name,
            "task_count": len(tasks),
            "fixed_count": fixed_count,
            "tasks": tasks,
            "generated_at": now_str,
            "inventory": args.inventory
        }
        all_idps.append(idp_record)

        # 写回 talent_profiles（存储IDP草稿）
        profiles["talents"][name].setdefault("idp", {})[args.inventory] = {
            "generated_at": now_str,
            "mentor": mentor_name,
            "goal": goal,
            "weak_dims": weak_dims,
            "tasks": tasks
        }

    save_profiles(profiles)
    write_excel(all_idps, Path(args.output))

    print(json.dumps({
        "status": "ok",
        "inventory": args.inventory,
        "generated_count": len(all_idps),
        "people": [idp["name"] for idp in all_idps],
        "output": args.output,
        "next_step": "请查看Excel草稿，确认后可发给相关人员；通过 update_progress.py 更新任务进展"
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
