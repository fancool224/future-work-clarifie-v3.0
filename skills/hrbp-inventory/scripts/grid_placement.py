#!/usr/bin/env python3
"""
25格落位计算 v2.0
- 盘点单元 = 一级部门 × 层级（各单元独立出矩阵）
- 绩效分档：方案C（相对位次为主，各盘点单元可覆写绝对阈值）
  - 相对位次：Top 20% = 优秀, 中 60% = 合格, 末 20% = 待改进
  - 绝对阈值覆写：config.performance_overrides[unit_key] = {"优秀": 1.1, "合格": 0.8}
- 输出：一个Excel，每个盘点单元一张矩阵Sheet + 汇总Sheet

用法:
  python3 scripts/grid_placement.py --inventory 2026H1 \
    --kpi_inputs <KPI文件1.xlsx>,<KPI文件2.xlsx> \
    --output artifacts/2026H1-盘点结果.xlsx \
    --data_dir data/inventories
"""
import argparse, json, sys
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))
from data_loader import load_kpi_file


# 梯队颜色
TIER_COLORS = {
    "第一梯队":   "1F4E79",
    "第二梯队":   "2E75B6",
    "骨干员工":   "70AD47",
    "待发展员工":  "FFD966",
    "待改进员工":  "F4B942",
    "问题员工":   "C00000",
}
TIER_FONT_COLORS = {
    "第一梯队":   "FFFFFF",
    "第二梯队":   "FFFFFF",
    "骨干员工":   "FFFFFF",
    "待发展员工":  "000000",
    "待改进员工":  "000000",
    "问题员工":   "FFFFFF",
}
HEADER_COLOR = "404040"
WHITE = "FFFFFF"
LIGHT_BLUE = "BDD7EE"


def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def set_header(cell, value, fg=HEADER_COLOR, fc=WHITE, bold=True, size=10):
    cell.value = value
    cell.font = Font(bold=bold, color=fc, size=size, name="微软雅黑")
    cell.fill = PatternFill("solid", fgColor=fg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()


def set_data(cell, value, bold=False, bg=None, fc="000000", align="left", size=9):
    cell.value = value
    cell.font = Font(bold=bold, size=size, color=fc, name="微软雅黑")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = thin_border()


def load_config(inventory: str, data_dir: str) -> dict:
    p = Path(data_dir) / inventory / "config.json"
    if not p.exists():
        print(json.dumps({"error": f"配置文件不存在：{p}"}, ensure_ascii=False))
        sys.exit(1)
    return json.loads(p.read_text(encoding="utf-8"))


def load_scores(inventory: str, data_dir: str) -> pd.DataFrame:
    p = Path(data_dir) / inventory / "scores_processed.json"
    if not p.exists():
        print(json.dumps({"error": f"评分处理文件不存在：{p}"}, ensure_ascii=False))
        sys.exit(1)
    data = json.loads(p.read_text(encoding="utf-8"))
    return pd.DataFrame(data["employees"]), data


def classify_performance_rank(values: pd.Series, top_pct=0.2, bottom_pct=0.2):
    """方案C默认：相对位次分档"""
    n = len(values)
    ranks = values.rank(method="average", ascending=True)
    percentiles = (ranks - 1) / max(n - 1, 1)
    result = []
    for p in percentiles:
        if p >= (1 - top_pct):
            result.append("优秀")
        elif p <= bottom_pct:
            result.append("待改进")
        else:
            result.append("合格")
    return result


def classify_performance_absolute(values: pd.Series, thresholds: dict):
    """方案C覆写：绝对阈值分档"""
    excellent = thresholds.get("优秀", 1.1)
    qualified = thresholds.get("合格", 0.8)
    result = []
    for v in values:
        if pd.isna(v):
            result.append(None)
        elif v >= excellent:
            result.append("优秀")
        elif v >= qualified:
            result.append("合格")
        else:
            result.append("待改进")
    return result


def classify_competency(scores: pd.Series, thresholds: dict):
    """素质轴：按绝对阈值（因为已经调平过，可比）"""
    excellent = thresholds.get("优秀", 110)
    good = thresholds.get("良好", 90)
    qualified = thresholds.get("合格", 75)
    result = []
    for s in scores:
        if pd.isna(s):
            result.append(None)
        elif s >= excellent:
            result.append("优秀")
        elif s >= good:
            result.append("良好")
        elif s >= qualified:
            result.append("合格")
        else:
            result.append("待改进")
    return result


# 25格落位规则（行=素质，列=绩效，各5档）
# 绩效列：1=待改进, 2=合格, 3=优秀 (简化为3档对应5列)
# 素质行：1=待改进, 2=合格, 3=良好, 4=优秀 (4档对应5行)
# 完整5×5 grid：
#   绩效 → 列 1-5（低到高）
#   素质 → 行 1-5（低到高）
#   格子编号 = (行-1)*5 + 列
PERF_TO_COL = {"待改进": 1, "合格": 3, "优秀": 5}  # 3档映射到1/3/5列
COMP_TO_ROW = {"待改进": 1, "合格": 2, "良好": 3, "优秀": 5}  # 4档映射到行

TIER_RULES = [
    ([25, 24, 23, 20], "第一梯队"),
    ([22, 21, 19, 18, 15], "第二梯队"),
    ([17, 14, 13, 16, 11, 10], "骨干员工"),
    ([12, 9, 8, 7, 6], "待发展员工"),
    ([5, 4, 3], "待改进员工"),
    ([2, 1], "问题员工"),
]


def grid_id(perf_cat, comp_cat):
    col = PERF_TO_COL.get(perf_cat, 3)
    row = COMP_TO_ROW.get(comp_cat, 2)
    return (row - 1) * 5 + col


def tier_for_grid(gid):
    for grids, tier in TIER_RULES:
        if gid in grids:
            return tier
    return "待发展员工"


def build_matrix_sheet(wb, sheet_name, unit_df, unit_key):
    """为一个盘点单元创建25格矩阵Sheet"""
    ws = wb.create_sheet(title=sheet_name[:31])

    # 标题行
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"人才盘点矩阵 — {unit_key}"
    title_cell.font = Font(bold=True, size=14, color="1F4E79", name="微软雅黑")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # 轴标签行
    ws["A3"] = "↑ 素质"
    ws["A3"].font = Font(bold=True, size=9, color="595959", name="微软雅黑")

    # 构建 5×5 grid
    perf_labels = ["绩效\n待改进", "绩效\n合格-", "绩效\n合格", "绩效\n合格+", "绩效\n优秀"]
    comp_labels = ["素质优秀", "素质良好", "素质合格+", "素质合格", "素质待改进"]

    # 表头（列）
    for ci, label in enumerate(perf_labels):
        col = ci + 2
        set_header(ws.cell(row=3, column=col), label, fg=LIGHT_BLUE, fc="1F4E79", size=8)
        ws.column_dimensions[get_column_letter(col)].width = 16

    ws.column_dimensions["A"].width = 12
    ws.row_dimensions[3].height = 36

    # 格子内容
    for ri in range(5):
        row_num = ri + 4
        comp_label = comp_labels[ri]
        ws.cell(row=row_num, column=1).value = comp_label
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=8, color="1F4E79", name="微软雅黑")
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row_num].height = 60

        for ci in range(5):
            col_num = ci + 2
            # grid_id 从左下到右上，行5=素质最高，行1=素质最低
            # 矩阵显示：row 4 = comp_row 5（优秀）, row 8 = comp_row 1（待改进）
            comp_row = 5 - ri
            perf_col = ci + 1
            gid = (comp_row - 1) * 5 + perf_col
            tier = tier_for_grid(gid)

            # 找该格子的员工
            members = unit_df[unit_df["grid_id"] == gid]
            names = "\n".join(members["name"].tolist()) if len(members) > 0 else ""

            cell = ws.cell(row=row_num, column=col_num)
            cell.value = f"格{gid}\n{names}" if names else f"格{gid}"
            cell.font = Font(size=8, color=TIER_FONT_COLORS.get(tier, "000000"), name="微软雅黑")
            cell.fill = PatternFill("solid", fgColor=TIER_COLORS.get(tier, "EEEEEE"))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border()

    # 图例
    legend_row = 10
    ws.cell(row=legend_row, column=1).value = "梯队图例："
    ws.cell(row=legend_row, column=1).font = Font(bold=True, size=9, name="微软雅黑")
    for i, (tier, color) in enumerate(TIER_COLORS.items()):
        c = ws.cell(row=legend_row, column=i + 2)
        c.value = tier
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(color=TIER_FONT_COLORS.get(tier, "000000"), size=9, name="微软雅黑")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

    return ws


def build_detail_sheet(wb, all_df):
    """员工明细 Sheet"""
    ws = wb.create_sheet(title="员工明细")
    headers = ["盘点单元", "姓名", "职位", "一级部门", "层级",
               "绩效达成率", "绩效分类", "绩效分档方式",
               "素质原始分", "素质调平分", "调平系数δ",
               "绩效轴(列)", "素质轴(行)", "格子编号", "梯队"]
    for ci, h in enumerate(headers, 1):
        set_header(ws.cell(row=1, column=ci), h, size=9)
    ws.row_dimensions[1].height = 36

    for ri, row in all_df.iterrows():
        r = ri + 2
        data = [
            row.get("unit_key", ""),
            row.get("name", ""),
            row.get("pos", ""),
            row.get("dept_l1", ""),
            row.get("level", ""),
            f"{row.get('kpi_rate', 0)*100:.1f}%" if pd.notna(row.get("kpi_rate")) else "—",
            row.get("perf_cat", ""),
            row.get("perf_method", ""),
            row.get("score_raw", ""),
            row.get("score_adjusted", ""),
            row.get("delta", ""),
            row.get("perf_col", ""),
            row.get("comp_row", ""),
            row.get("grid_id", ""),
            row.get("tier", ""),
        ]
        for ci, val in enumerate(data, 1):
            tier = row.get("tier", "")
            bg = TIER_COLORS.get(tier) if ci == 15 else None
            fc = TIER_FONT_COLORS.get(tier, "000000") if ci == 15 else "000000"
            set_data(ws.cell(row=r, column=ci), val, bg=bg, fc=fc)

    for i, w in enumerate([18, 8, 14, 10, 8, 10, 8, 10, 10, 10, 8, 8, 8, 8, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_tier_sheet(wb, all_df):
    """梯队名单 Sheet"""
    ws = wb.create_sheet(title="梯队名单")
    tier_order = ["第一梯队", "第二梯队", "骨干员工", "待发展员工", "待改进员工", "问题员工"]
    col = 1
    for tier in tier_order:
        members = all_df[all_df["tier"] == tier].sort_values(["dept_l1", "level", "name"])
        set_header(ws.cell(row=1, column=col), f"{tier} ({len(members)}人)",
                   fg=TIER_COLORS.get(tier, "999999"),
                   fc=TIER_FONT_COLORS.get(tier, "000000"))
        ws.column_dimensions[get_column_letter(col)].width = 16
        for ri, (_, row) in enumerate(members.iterrows(), 2):
            c = ws.cell(row=ri, column=col)
            c.value = f"{row['name']}({row['dept_l1']}-{row['level']})"
            c.font = Font(size=9, name="微软雅黑")
            c.border = thin_border()
        col += 1


def build_summary_sheet(wb, all_df, anomaly_raters, m_global, group_stats):
    """统计摘要 + 数据质量"""
    ws = wb.create_sheet(title="统计摘要")
    r = 1

    # 整体分布
    ws.cell(row=r, column=1).value = "整体梯队分布"
    ws.cell(row=r, column=1).font = Font(bold=True, size=11, name="微软雅黑")
    r += 1
    tier_order = ["第一梯队", "第二梯队", "骨干员工", "待发展员工", "待改进员工", "问题员工"]
    for h in ["梯队", "人数", "占比"]:
        set_header(ws.cell(row=r, column=tier_order.index(h) + 1 if h in tier_order else ["梯队","人数","占比"].index(h)+1), h, size=9)
    r += 1
    total = len(all_df)
    for tier in tier_order:
        n = len(all_df[all_df["tier"] == tier])
        ws.cell(row=r, column=1).value = tier
        ws.cell(row=r, column=2).value = n
        ws.cell(row=r, column=3).value = f"{n/total*100:.1f}%"
        r += 1

    r += 1
    # 调平结果
    ws.cell(row=r, column=1).value = f"全局基准 M_global = {m_global:.2f}"
    ws.cell(row=r, column=1).font = Font(bold=True, size=10, name="微软雅黑")
    r += 1
    for gs in group_stats:
        ws.cell(row=r, column=1).value = f"{gs['dept']}-{gs['level']}"
        ws.cell(row=r, column=2).value = f"M_g={gs['M_g']:.2f}"
        ws.cell(row=r, column=3).value = f"δ={gs['delta']:+.2f}"
        ws.cell(row=r, column=4).value = f"有效{gs['n_valid']}/{gs['n_total']}条"
        if gs.get("fallback"):
            ws.cell(row=r, column=5).value = "⚠️ FALLBACK"
            ws.cell(row=r, column=5).font = Font(color="C00000", bold=True, name="微软雅黑")
        r += 1

    r += 1
    # 异常评分者
    ws.cell(row=r, column=1).value = f"异常评分者 ({len(anomaly_raters)}名)"
    ws.cell(row=r, column=1).font = Font(bold=True, size=10, color="C00000", name="微软雅黑")
    r += 1
    for ar in anomaly_raters:
        ws.cell(row=r, column=1).value = ar["rater"]
        ws.cell(row=r, column=2).value = ar["level"]
        ws.cell(row=r, column=3).value = ar["flag"]
        ws.cell(row=r, column=4).value = ar["reason"]
        r += 1

    for i in range(1, 6):
        ws.column_dimensions[get_column_letter(i)].width = 20


def cmd_full(args):
    config = load_config(args.inventory, args.data_dir)
    scores_df, scores_meta = load_scores(args.inventory, args.data_dir)
    anomaly_raters = scores_meta.get("anomaly_raters", [])
    m_global = scores_meta.get("M_global", 0)
    group_stats = scores_meta.get("group_stats", [])

    # 加载 KPI 数据（支持多文件）
    kpi_files = [f.strip() for f in args.kpi_inputs.split(",") if f.strip()]
    kpi_frames = []
    for f in kpi_files:
        try:
            kf = load_kpi_file(f, config)
            kpi_frames.append(kf)
        except Exception as e:
            print(f"⚠️ 跳过 KPI 文件 {f}：{e}", file=sys.stderr)
    if not kpi_frames:
        print(json.dumps({"error": "没有可用的KPI文件"}, ensure_ascii=False))
        sys.exit(1)
    kpi_df = pd.concat(kpi_frames, ignore_index=True).drop_duplicates("unique_key")

    # 合并素质 + KPI
    merged = scores_df.merge(kpi_df[["unique_key", "kpi_rate", "kpi_type"]], on="unique_key", how="left")

    # 绩效分档（方案C：相对位次为主，可按盘点单元覆写）
    perf_overrides = config.get("performance_overrides", {})
    all_rows = []

    for (dept, level), grp in merged.groupby(["dept_l1", "level"]):
        grp = grp.copy()
        unit_key = f"{dept}-{level}"

        # 判断是否有绝对阈值覆写
        override = perf_overrides.get(unit_key) or perf_overrides.get(dept)
        if override:
            grp["perf_cat"] = classify_performance_absolute(grp["kpi_rate"], override)
            grp["perf_method"] = "绝对阈值"
        else:
            grp["perf_cat"] = classify_performance_rank(grp["kpi_rate"])
            grp["perf_method"] = "相对位次"

        # 素质分档（绝对阈值，已调平可比）
        comp_thresholds = config.get("competency_thresholds", {"优秀": 110, "良好": 90, "合格": 75})
        grp["comp_cat"] = classify_competency(grp["score_adjusted"], comp_thresholds)

        grp["unit_key"] = unit_key
        grp["perf_col"] = grp["perf_cat"].map(PERF_TO_COL).fillna(3).astype(int)
        grp["comp_row"] = grp["comp_cat"].map(COMP_TO_ROW).fillna(2).astype(int)
        grp["grid_id"] = grp.apply(lambda r: grid_id(r["perf_cat"], r["comp_cat"]), axis=1)
        grp["tier"] = grp["grid_id"].apply(tier_for_grid)

        all_rows.append(grp)

    all_df = pd.concat(all_rows, ignore_index=True)

    # 构建 Excel
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认 Sheet

    # 各盘点单元矩阵
    units = sorted(all_df["unit_key"].unique())
    for unit_key in units:
        unit_df = all_df[all_df["unit_key"] == unit_key]
        sheet_name = unit_key.replace("/", "-")
        build_matrix_sheet(wb, sheet_name, unit_df, unit_key)

    build_detail_sheet(wb, all_df)
    build_tier_sheet(wb, all_df)
    build_summary_sheet(wb, all_df, anomaly_raters, m_global, group_stats)

    wb.save(output_path)

    # 梯队分布统计
    tier_dist = all_df["tier"].value_counts().to_dict()
    missing = all_df[all_df["kpi_rate"].isna()]

    result = {
        "status": "ok",
        "inventory": args.inventory,
        "output": str(output_path),
        "total_employees": len(all_df),
        "inventory_units": units,
        "tier_distribution": tier_dist,
        "missing_kpi": len(missing),
        "missing_kpi_names": missing["name"].tolist() if len(missing) > 0 else [],
        "anomaly_raters": anomaly_raters,
        "perf_method_used": {k: v["perf_method"].iloc[0] for k, v in all_df.groupby("unit_key") if "perf_method" in v.columns},
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


def main():
    parser = argparse.ArgumentParser(description="25格落位 v2.0（盘点单元 = 部门×层级）")
    parser.add_argument("--inventory", required=True)
    parser.add_argument("--kpi_inputs", required=True, help="KPI文件路径，多个用逗号分隔")
    parser.add_argument("--output", required=True)
    parser.add_argument("--data_dir", default="data/inventories")
    args = parser.parse_args()
    cmd_full(args)


if __name__ == "__main__":
    main()
