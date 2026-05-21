#!/usr/bin/env python3
"""
导出胜任力模型为 Excel
用法: uv run --python 3.12 scripts/export_model.py --model <tangjiu|custom_json_path> --output <路径> [--layer <管理层|员工层|全部>]
"""
import argparse, json, sys
from pathlib import Path

SKILL_DIR = Path(__file__).parent.parent
MODEL_PATH = SKILL_DIR / "references" / "tangjiu_model.json"

def load_model(model_arg):
    if model_arg == "tangjiu":
        with open(MODEL_PATH, encoding="utf-8") as f:
            return json.load(f)
    else:
        p = Path(model_arg)
        if not p.exists():
            print(json.dumps({"error": f"文件不存在: {model_arg}"}, ensure_ascii=False))
            sys.exit(1)
        with open(p, encoding="utf-8") as f:
            return json.load(f)

def export_excel(model, output_path, layer_filter="全部"):
    try:
        import sys
        sys.path.insert(0, str(SKILL_DIR.parent.parent / "deps" / "python"))
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(json.dumps({"error": "缺少 openpyxl，请先安装: uv pip install --target deps/python/ openpyxl"}, ensure_ascii=False))
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "胜任力模型"

    # 颜色定义
    HEADER_COLOR = "1F4E79"
    DIM_COLOR = "2E75B6"
    ELEM_COLOR = "BDD7EE"
    ALT_ROW = "F2F7FB"
    WHITE = "FFFFFF"

    def make_border():
        thin = Side(style="thin", color="BFBFBF")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_cell(cell, value, bold=True, color=HEADER_COLOR, font_color="FFFFFF", size=10):
        cell.value = value
        cell.font = Font(bold=bold, color=font_color, size=size)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border()

    def data_cell(cell, value, bold=False, bg=None, indent=0):
        cell.value = value
        cell.font = Font(bold=bold, size=10)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=indent)
        cell.border = make_border()

    # 标题行
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"{model['model_name']} — 能力词典"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 列标题
    headers = ["维度", "要素", "要素定义", "行为指标", "管理层序号", "员工层序号", "适用层级", "评分参考(A/B/C/D/E)"]
    col_widths = [12, 12, 35, 55, 10, 10, 15, 45]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col)
        header_cell(cell, h)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    scoring = model["scoring"]
    score_str = "  ".join([f"{k}={v['label']}({v['score']}分)" for k, v in scoring.items()])

    row_num = 3
    for dim in model["dimensions"]:
        for elem in dim["elements"]:
            for ind in elem["indicators"]:
                # 层级过滤
                if layer_filter == "管理层" and ind["mgmt_seq"] is None:
                    continue
                if layer_filter == "员工层" and ind["staff_seq"] is None:
                    continue

                layers = []
                if ind["mgmt_seq"]:
                    layers.append("管理层")
                if ind["staff_seq"]:
                    layers.append("员工层")
                layer_str = "、".join(layers)

                bg = ALT_ROW if row_num % 2 == 0 else WHITE

                data_cell(ws.cell(row=row_num, column=1), dim["name"], bold=True, bg=ELEM_COLOR)
                data_cell(ws.cell(row=row_num, column=2), elem["name"], bold=True, bg=ELEM_COLOR)
                data_cell(ws.cell(row=row_num, column=3), elem["definition"], bg=bg)
                data_cell(ws.cell(row=row_num, column=4), ind["behavior"], bg=bg)
                data_cell(ws.cell(row=row_num, column=5), ind["mgmt_seq"] or "", bg=bg)
                data_cell(ws.cell(row=row_num, column=6), ind["staff_seq"] or "", bg=bg)
                data_cell(ws.cell(row=row_num, column=7), layer_str, bg=bg)
                data_cell(ws.cell(row=row_num, column=8), score_str, bg=bg)

                ws.row_dimensions[row_num].height = 50
                row_num += 1

    # 评分说明 sheet
    ws2 = wb.create_sheet("评分标准")
    ws2["A1"].value = "能力评估评分标准"
    ws2["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor=HEADER_COLOR)
    ws2.merge_cells("A1:D1")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 25

    h2_row = ["量级", "标签", "分数", "阐释"]
    for col, h in enumerate(h2_row, 1):
        c = ws2.cell(row=2, column=col)
        header_cell(c, h)

    for i, (k, v) in enumerate(scoring.items(), 3):
        ws2.cell(row=i, column=1).value = k
        ws2.cell(row=i, column=2).value = v["label"]
        ws2.cell(row=i, column=3).value = v["score"]
        ws2.cell(row=i, column=4).value = v["desc"]
        for col in range(1, 5):
            ws2.cell(row=i, column=col).border = make_border()
            ws2.cell(row=i, column=col).alignment = Alignment(vertical="center", wrap_text=True)

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 8
    ws2.column_dimensions["D"].width = 60

    ws2["A8"].value = "⚠️ 注意：评分非线性，A与B之间差距为2分，体现对高频优秀行为的刻意放大。"
    ws2.merge_cells("A8:D8")
    ws2["A8"].font = Font(italic=True, color="C00000")

    wb.save(output_path)
    print(json.dumps({"status": "ok", "output": str(output_path), "rows": row_num - 3}, ensure_ascii=False))

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--model", default="tangjiu", help="tangjiu 或自定义JSON路径")
    parser.add_argument("--output", required=True, help="输出Excel路径")
    parser.add_argument("--layer", choices=["管理层", "员工层", "全部"], default="全部")
    args = parser.parse_args()

    model = load_model(args.model)
    export_excel(model, args.output, layer_filter=args.layer)

if __name__ == "__main__":
    main()
